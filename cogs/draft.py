import csv
import json
import os
from pathlib import Path
from datetime import datetime, timezone

import discord
from discord import app_commands
from discord.ext import commands, tasks
from openpyxl import Workbook
from openpyxl.styles import Font

from espn_api.football.constant import POSITION_MAP

from database.database import (
    accept_trade_proposal,
    cancel_trade_proposal,
    claim_team,
    create_trade_proposal,
    decline_trade_proposal,
    end_draft,
    expire_current_pick_and_advance,
    get_all_draft_picks,
    get_all_pick_clocks,
    get_all_expired_picks,
    get_all_pick_ownership,
    get_all_pick_trades,
    get_all_team_claims,
    get_current_draft_team,
    get_current_pick_clock,
    get_drafted_player_ids,
    get_due_clock_notifications,
    get_draft_order,
    get_draft_state,
    get_pending_trade_proposals,
    get_pending_trade_proposals_for_user,
    get_pick_owner,
    get_oldest_expired_pick_for_team,
    get_outstanding_expired_pick_count,
    get_team_claim_by_team,
    get_team_claim_by_user,
    get_trade_proposal,
    invalidate_pending_trade_proposals_for_picks,
    mark_clock_notification_sent,
    remove_team_claim,
    save_draft_channel,
    save_draft_pick,
    save_draftboard_message,
    save_trade_proposal_message,
    start_new_draft,
    trade_draft_picks,
    undo_last_draft_pick,
    undo_last_pick_trade,
    update_trade_proposal_status,
)

from services.espn_service import (
    get_league,
    normalize_team_name,
)


# =============================================================
# SHARED CONSTANTS
# =============================================================

ALLOWED_POSITIONS = {
    "QB",
    "RB",
    "WR",
    "TE",
    "K",
    "D/ST",
    "DB",
    "LB",
}

OWNER_ROLE_NAME = "Owner"

DRAFT_ADMIN_ROLES = {
    "Commissioner",
    "Developer",
}


# =============================================================
# TRADE PROPOSAL VIEW
# =============================================================

class TradeProposalView(discord.ui.View):

    def __init__(
        self,
        draft_cog,
        proposal_id,
        recipient_user_id,
    ):
        super().__init__(
            timeout=None
        )

        self.draft_cog = draft_cog
        self.proposal_id = proposal_id
        self.recipient_user_id = recipient_user_id

        accept_button = discord.ui.Button(
            label="Accept Trade",
            style=discord.ButtonStyle.success,
            custom_id=(
                f"draft_trade_accept:"
                f"{proposal_id}"
            ),
        )

        decline_button = discord.ui.Button(
            label="Decline Trade",
            style=discord.ButtonStyle.danger,
            custom_id=(
                f"draft_trade_decline:"
                f"{proposal_id}"
            ),
        )

        accept_button.callback = (
            self.accept_trade
        )

        decline_button.callback = (
            self.decline_trade
        )

        self.add_item(
            accept_button
        )

        self.add_item(
            decline_button
        )

    async def interaction_check(
        self,
        interaction: discord.Interaction,
    ):
        if (
            interaction.user.id
            != self.recipient_user_id
        ):
            await interaction.response.send_message(
                (
                    "❌ Only the manager receiving "
                    "this trade proposal can "
                    "accept or decline it."
                ),
                ephemeral=True,
            )

            return False

        return True

    def disable_all_buttons(
        self,
    ):
        for item in self.children:
            item.disabled = True

    async def accept_trade(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer()

        try:
            proposal = get_trade_proposal(
                self.proposal_id
            )

            if proposal is None:
                self.disable_all_buttons()

                await interaction.edit_original_response(
                    content=(
                        "❌ This trade proposal "
                        "no longer exists."
                    ),
                    view=self,
                )
                return

            if (
                proposal["status"]
                != "PENDING"
            ):
                self.disable_all_buttons()

                await interaction.edit_original_response(
                    content=(
                        "❌ This trade proposal is "
                        f"already **{proposal['status']}**."
                    ),
                    view=self,
                )
                return

            result = accept_trade_proposal(
                proposal_id=(
                    self.proposal_id
                ),
                accepting_discord_user_id=(
                    interaction.user.id
                ),
            )

            first = result[
                "pick_a"
            ]

            second = result[
                "pick_b"
            ]

            self.disable_all_buttons()

            embed = discord.Embed(
                title=(
                    "✅ Draft Pick Trade Accepted"
                ),
                description=(
                    "The pick swap is complete."
                ),
            )

            embed.add_field(
                name=(
                    f"Overall "
                    f"{first['overall_pick']} "
                    f"— Round "
                    f"{first['round_number']} "
                    f"Pick "
                    f"{first['pick_in_round']}"
                ),
                value=(
                    f"~~{first['previous_team_name']}~~\n"
                    f"➡️ **{first['new_team_name']}**"
                ),
                inline=False,
            )

            embed.add_field(
                name=(
                    f"Overall "
                    f"{second['overall_pick']} "
                    f"— Round "
                    f"{second['round_number']} "
                    f"Pick "
                    f"{second['pick_in_round']}"
                ),
                value=(
                    f"~~{second['previous_team_name']}~~\n"
                    f"➡️ **{second['new_team_name']}**"
                ),
                inline=False,
            )

            await interaction.edit_original_response(
                content=None,
                embed=embed,
                view=self,
            )

            await self.draft_cog.refresh_draftboard()

        except ValueError as error:

            proposal = get_trade_proposal(
                self.proposal_id
            )

            if (
                proposal is not None
                and proposal["status"]
                == "PENDING"
            ):
                update_trade_proposal_status(
                    self.proposal_id,
                    "INVALID",
                )

            self.disable_all_buttons()

            await interaction.edit_original_response(
                content=(
                    "❌ This trade can no longer "
                    f"be completed.\n\n{error}"
                ),
                embed=None,
                view=self,
            )

        except Exception as error:
            print(
                (
                    "Accept Trade Proposal Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "accepting the trade."
                ),
                ephemeral=True,
            )

    async def decline_trade(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer()

        try:
            proposal = decline_trade_proposal(
                proposal_id=(
                    self.proposal_id
                ),
                declining_discord_user_id=(
                    interaction.user.id
                ),
            )

            self.disable_all_buttons()

            embed = discord.Embed(
                title=(
                    "❌ Draft Pick Trade Declined"
                ),
                description=(
                    f"**{proposal['recipient_team_name']}** "
                    "declined the proposal from "
                    f"**{proposal['proposer_team_name']}**."
                ),
            )

            await interaction.edit_original_response(
                content=None,
                embed=embed,
                view=self,
            )

        except ValueError as error:

            self.disable_all_buttons()

            await interaction.edit_original_response(
                content=(
                    f"❌ {error}"
                ),
                embed=None,
                view=self,
            )

        except Exception as error:
            print(
                (
                    "Decline Trade Proposal Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "declining the trade."
                ),
                ephemeral=True,
            )


# =============================================================
# DRAFT ORDER MODAL
# =============================================================

class DraftOrderModal(discord.ui.Modal):

    def __init__(
        self,
        draft_cog,
        total_rounds,
    ):
        super().__init__(
            title="Set Draft Order",
            timeout=300,
        )

        self.draft_cog = draft_cog
        self.total_rounds = total_rounds

        self.draft_order = discord.ui.TextInput(
            label="Draft Order",
            style=discord.TextStyle.paragraph,
            placeholder=(
                "Enter one fantasy team per line.\n\n"
                "Example:\n"
                "Future WWE Star\n"
                "Dallas Cowgirls\n"
                "Jungle Jeopardy"
            ),
            required=True,
            max_length=2000,
        )

        self.add_item(
            self.draft_order
        )

    async def on_submit(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            league = get_league()

            submitted_names = [
                line.strip()
                for line
                in self.draft_order.value.splitlines()
                if line.strip()
            ]

            expected_team_count = len(
                league.teams
            )

            if (
                len(submitted_names)
                != expected_team_count
            ):
                await interaction.followup.send(
                    (
                        f"❌ The league has "
                        f"{expected_team_count} teams, "
                        f"but you entered "
                        f"{len(submitted_names)}.\n\n"
                        "Please run `/start-draft` again "
                        "and enter every team exactly once."
                    ),
                    ephemeral=True,
                )
                return

            team_lookup = {
                normalize_team_name(
                    team.team_name
                ): team
                for team in league.teams
            }

            draft_order = []
            seen_team_ids = set()

            for submitted_name in submitted_names:

                team = team_lookup.get(
                    normalize_team_name(
                        submitted_name
                    )
                )

                if team is None:
                    await interaction.followup.send(
                        (
                            "❌ Could not find the ESPN team "
                            f'**"{submitted_name}"**.\n\n'
                            "Please check the spelling "
                            "and try again."
                        ),
                        ephemeral=True,
                    )
                    return

                if (
                    team.team_id
                    in seen_team_ids
                ):
                    await interaction.followup.send(
                        (
                            f"❌ **{team.team_name}** "
                            "appears more than once "
                            "in the draft order."
                        ),
                        ephemeral=True,
                    )
                    return

                seen_team_ids.add(
                    team.team_id
                )

                draft_order.append(
                    {
                        "espn_team_id": (
                            team.team_id
                        ),
                        "team_name": (
                            team.team_name
                        ),
                    }
                )

            start_new_draft(
                draft_order,
                self.total_rounds,
            )

            save_draft_channel(
                interaction.channel.id
            )

            order_text = ""

            for position, team in enumerate(
                draft_order,
                start=1,
            ):
                order_text += (
                    f"**{position}.** "
                    f"{team['team_name']}\n"
                )

            total_picks = (
                len(draft_order)
                * self.total_rounds
            )

            embed = discord.Embed(
                title=(
                    "🏈 Flint Michigan Megabowl "
                    "Draft Started"
                ),
                description=(
                    "Snake Draft • "
                    f"{self.total_rounds} Rounds • "
                    f"{len(draft_order)} Teams • "
                    f"{total_picks} Picks"
                ),
            )

            embed.add_field(
                name="Draft Order",
                value=order_text,
                inline=False,
            )

            embed.add_field(
                name="⏰ On the Clock",
                value=(
                    "**Pick 1 — "
                    f"{draft_order[0]['team_name']}**"
                ),
                inline=False,
            )

            await interaction.channel.send(
                embed=embed
            )

            draftboard_message = (
                await self.draft_cog.create_draftboard_message()
            )

            await self.draft_cog.process_draft_clock()

            await interaction.followup.send(
                (
                    "✅ Draft initialized successfully.\n"
                    f"Draft board created in "
                    f"{draftboard_message.channel.mention}."
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Start Draft Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "starting the draft."
                ),
                ephemeral=True,
            )


# =============================================================
# END DRAFT CONFIRMATION
# =============================================================

class EndDraftConfirmationView(
    discord.ui.View
):

    def __init__(
        self,
        draft_cog,
        requesting_user_id,
    ):
        super().__init__(
            timeout=60
        )

        self.draft_cog = draft_cog
        self.requesting_user_id = (
            requesting_user_id
        )

    async def interaction_check(
        self,
        interaction: discord.Interaction,
    ):
        if (
            interaction.user.id
            != self.requesting_user_id
        ):
            await interaction.response.send_message(
                (
                    "❌ Only the person who requested "
                    "the draft end can confirm it."
                ),
                ephemeral=True,
            )

            return False

        return True

    @discord.ui.button(
        label="End Draft",
        style=discord.ButtonStyle.danger,
    )
    async def confirm(
        self,
        interaction: discord.Interaction,
        button: discord.ui.Button,
    ):
        end_draft()

        for item in self.children:
            item.disabled = True

        await interaction.response.edit_message(
            content=(
                "✅ The draft has been ended."
            ),
            view=self,
        )

        await self.draft_cog.refresh_draftboard()

        self.stop()

    @discord.ui.button(
        label="Cancel",
        style=discord.ButtonStyle.secondary,
    )
    async def cancel(
        self,
        interaction: discord.Interaction,
        button: discord.ui.Button,
    ):
        for item in self.children:
            item.disabled = True

        await interaction.response.edit_message(
            content=(
                "Draft end cancelled."
            ),
            view=self,
        )

        self.stop()


# =============================================================
# DRAFT COG
# =============================================================

class DraftCog(commands.Cog):

    def __init__(
        self,
        bot,
    ):
        self.bot = bot
        self.draft_clock_loop.start()

    def cog_unload(
        self,
    ):
        self.draft_clock_loop.cancel()

    async def cog_load(
        self,
    ):
        try:
            pending_proposals = (
                get_pending_trade_proposals()
            )

            for proposal in pending_proposals:

                self.bot.add_view(
                    TradeProposalView(
                        draft_cog=self,
                        proposal_id=(
                            proposal["id"]
                        ),
                        recipient_user_id=(
                            proposal[
                                "recipient_discord_user_id"
                            ]
                        ),
                    )
                )

            if pending_proposals:
                print(
                    (
                        "Restored "
                        f"{len(pending_proposals)} "
                        "pending trade proposal view(s)."
                    )
                )

        except Exception as error:
            print(
                (
                    "Trade Proposal View Restore Error: "
                    f"{error}"
                )
            )

    # =========================================================
    # DRAFT CLOCK
    # =========================================================

    def _parse_clock_timestamp(
        self,
        value,
    ):
        if not value:
            return None

        parsed = datetime.fromisoformat(
            str(value).replace("Z", "+00:00")
        )

        if parsed.tzinfo is None:
            parsed = parsed.replace(
                tzinfo=timezone.utc
            )

        return parsed.astimezone(
            timezone.utc
        )

    async def get_draft_channel(
        self,
    ):
        state = get_draft_state()
        channel_id = state[
            "draft_channel_id"
        ]

        if not channel_id:
            return None

        channel = self.bot.get_channel(
            channel_id
        )

        if channel is None:
            try:
                channel = await self.bot.fetch_channel(
                    channel_id
                )
            except discord.DiscordException:
                return None

        return channel

    def get_pick_owner_mention(
        self,
        ownership,
    ):
        if ownership is None:
            return None

        claim = get_team_claim_by_team(
            ownership[
                "current_espn_team_id"
            ]
        )

        if claim is None:
            return None

        return (
            f"<@{claim['discord_user_id']}>"
        )

    def format_clock_deadline(
        self,
        clock,
    ):
        if clock is None:
            return None

        expires_at = self._parse_clock_timestamp(
            clock["clock_expires_at"]
        )

        if expires_at is None:
            return None

        timestamp = int(
            expires_at.timestamp()
        )

        return (
            f"<t:{timestamp}:F> "
            f"(<t:{timestamp}:R>)"
        )

    async def send_clock_notification(
        self,
        notification_type,
        ownership,
        clock,
    ):
        channel = await self.get_draft_channel()

        if channel is None or ownership is None:
            return False

        mention = self.get_pick_owner_mention(
            ownership
        )

        addressee = (
            mention
            if mention
            else f"**{ownership['current_team_name']}**"
        )

        pick_text = (
            f"Round {ownership['round_number']} • "
            f"Pick {ownership['pick_in_round']} • "
            f"Overall {ownership['overall_pick']}"
        )

        deadline = self.format_clock_deadline(
            clock
        )

        if notification_type == "START":
            message = (
                f"⏰ {addressee} is **ON THE CLOCK**.\n"
                f"{pick_text}\n"
                f"You have **12 hours** to make the pick."
            )

            if deadline:
                message += (
                    f"\nClock expires {deadline}."
                )

        elif notification_type == "SIX_HOUR":
            message = (
                f"⏳ {addressee} — **6 hours remain** "
                f"on your draft clock.\n{pick_text}"
            )

            if deadline:
                message += (
                    f"\nClock expires {deadline}."
                )

        elif notification_type == "THIRTY_MINUTE":
            message = (
                f"🚨 {addressee} — **30 minutes remain** "
                f"on your draft clock.\n{pick_text}"
            )

            if deadline:
                message += (
                    f"\nClock expires {deadline}."
                )

        else:
            return False

        await channel.send(
            message
        )

        return True

    async def process_draft_clock(
        self,
    ):
        state = get_draft_state()

        if not state["active"]:
            return

        due = get_due_clock_notifications()

        if not due:
            return

        notification_type = due[0]
        current_clock = get_current_pick_clock()

        if current_clock is None:
            return

        overall_pick = current_clock[
            "overall_pick"
        ]
        ownership = get_pick_owner(
            overall_pick
        )

        if notification_type == "EXPIRED":
            result = expire_current_pick_and_advance()

            if result is None:
                return

            channel = await self.get_draft_channel()
            expired_owner = result[
                "expired_owner"
            ]

            if channel is not None and expired_owner is not None:
                mention = self.get_pick_owner_mention(
                    expired_owner
                )
                addressee = (
                    mention
                    if mention
                    else f"**{expired_owner['current_team_name']}**"
                )

                await channel.send(
                    (
                        f"⌛ {addressee}'s clock expired for "
                        f"**Overall Pick {result['expired_pick']}**.\n"
                        "That pick is now a **catch-up pick** and "
                        "may be made at any time. The draft clock "
                        "has advanced to the next scheduled pick."
                    )
                )

            mark_clock_notification_sent(
                result["expired_pick"],
                "EXPIRED",
            )

            await self.refresh_draftboard()

            if result["next_pick"] is not None:
                next_owner = get_pick_owner(
                    result["next_pick"]
                )
                next_clock = result[
                    "next_clock"
                ]

                if await self.send_clock_notification(
                    "START",
                    next_owner,
                    next_clock,
                ):
                    mark_clock_notification_sent(
                        result["next_pick"],
                        "START",
                    )

            return

        if await self.send_clock_notification(
            notification_type,
            ownership,
            current_clock,
        ):
            mark_clock_notification_sent(
                overall_pick,
                notification_type,
            )

    @tasks.loop(
        seconds=30
    )
    async def draft_clock_loop(
        self,
    ):
        try:
            await self.process_draft_clock()
        except Exception as error:
            print(
                f"Draft Clock Error: {error}"
            )

    @draft_clock_loop.before_loop
    async def before_draft_clock_loop(
        self,
    ):
        await self.bot.wait_until_ready()

    # =========================================================
    # ROLE / PERMISSION HELPERS
    # =========================================================

    def has_draft_admin_role(
        self,
        interaction: discord.Interaction,
    ):
        if not isinstance(
            interaction.user,
            discord.Member,
        ):
            return False

        user_roles = {
            role.name
            for role
            in interaction.user.roles
        }

        return bool(
            DRAFT_ADMIN_ROLES
            & user_roles
        )

    def get_owner_role(
        self,
        guild: discord.Guild,
    ):
        role = discord.utils.get(
            guild.roles,
            name=OWNER_ROLE_NAME,
        )

        if role is None:
            raise ValueError(
                (
                    f'The Discord role "{OWNER_ROLE_NAME}" '
                    "could not be found."
                )
            )

        return role

    async def add_owner_role(
        self,
        member: discord.Member,
    ):
        owner_role = self.get_owner_role(
            member.guild
        )

        if owner_role in member.roles:
            return False

        await member.add_roles(
            owner_role,
            reason=(
                "Fantasy team successfully claimed."
            ),
        )

        return True

    async def remove_owner_role(
        self,
        member: discord.Member,
    ):
        owner_role = self.get_owner_role(
            member.guild
        )

        if owner_role not in member.roles:
            return False

        await member.remove_roles(
            owner_role,
            reason=(
                "Fantasy team claim removed."
            ),
        )

        return True

    async def get_guild_member(
        self,
        guild: discord.Guild,
        user_id: int,
    ):
        member = guild.get_member(
            user_id
        )

        if member is not None:
            return member

        try:
            return await guild.fetch_member(
                user_id
            )

        except discord.NotFound:
            return None

    # =========================================================
    # ESPN RANK HELPERS
    # =========================================================

    def get_kona_position_rank(
        self,
        player_data,
    ):
        rankings = player_data.get(
            "rankings",
            {}
        )

        for ranking_group in rankings.values():

            if not isinstance(
                ranking_group,
                list,
            ):
                continue

            for ranking in ranking_group:

                if not isinstance(
                    ranking,
                    dict,
                ):
                    continue

                if (
                    ranking.get(
                        "rankSourceId"
                    ) == 0
                    and ranking.get(
                        "rankType"
                    ) == "STANDARD"
                ):
                    average_rank = ranking.get(
                        "averageRank"
                    )

                    if isinstance(
                        average_rank,
                        (int, float),
                    ) and average_rank > 0:
                        return float(
                            average_rank
                        )

        return None

    def format_position_rank(
        self,
        position,
        rank,
    ):
        if rank is None:
            return "Unranked"

        rounded_rank = int(
            round(rank)
        )

        return (
            f"{position}{rounded_rank}"
        )

    # =========================================================
    # KONA PLAYER POOL
    # =========================================================

    def get_ranked_player_pool(
        self,
        league,
        drafted_player_ids,
        position_filter=None,
        search_text=None,
    ):
        player_filter = {
            "limit": 200,
            "sortDraftRanks": {
                "sortPriority": 1,
                "sortAsc": True,
                "value": "STANDARD",
            },
        }

        if position_filter:

            slot_id = POSITION_MAP.get(
                position_filter
            )

            if slot_id is None:
                return []

            player_filter[
                "filterSlotIds"
            ] = {
                "value": [
                    slot_id
                ]
            }

        if search_text:

            matching_ids = []

            for (
                player_name,
                player_id,
            ) in league.player_map.items():

                if not isinstance(
                    player_id,
                    int,
                ):
                    continue

                if (
                    player_id
                    in drafted_player_ids
                ):
                    continue

                if (
                    search_text
                    not in player_name.lower()
                ):
                    continue

                matching_ids.append(
                    player_id
                )

            if not matching_ids:
                return []

            matching_ids = (
                matching_ids[:100]
            )

            player_filter[
                "filterIds"
            ] = {
                "value": matching_ids
            }

            player_filter[
                "limit"
            ] = len(
                matching_ids
            )

        params = {
            "view": "kona_player_info",
            "scoringPeriodId": (
                league.current_week
            ),
        }

        filters = {
            "players": player_filter
        }

        headers = {
            "x-fantasy-filter": json.dumps(
                filters
            )
        }

        data = (
            league.espn_request.league_get(
                params=params,
                headers=headers,
            )
        )

        player_entries = data.get(
            "players",
            []
        )

        results = []

        for entry in player_entries:

            player_data = entry.get(
                "player",
                {}
            )

            player_id = player_data.get(
                "id"
            )

            if not isinstance(
                player_id,
                int,
            ):
                continue

            if (
                player_id
                in drafted_player_ids
            ):
                continue

            player_name = player_data.get(
                "fullName",
                "",
            )

            if (
                search_text
                and search_text
                not in player_name.lower()
            ):
                continue

            rank = (
                self.get_kona_position_rank(
                    player_data
                )
            )

            results.append(
                {
                    "player_id": (
                        player_id
                    ),
                    "name": (
                        player_name
                    ),
                    "rank": (
                        rank
                    ),
                    "raw": (
                        player_data
                    ),
                }
            )

        results.sort(
            key=lambda player: (
                (
                    player["rank"]
                    if player["rank"]
                    is not None
                    else 999999
                ),
                player["name"],
            )
        )

        return results

    # =========================================================
    # DRAFTBOARD CHANNEL
    # =========================================================

    async def get_draftboard_channel(
        self,
    ):
        channel_id = os.getenv(
            "DRAFTBOARD_CHANNEL_ID"
        )

        if not channel_id:
            raise ValueError(
                (
                    "DRAFTBOARD_CHANNEL_ID "
                    "is missing from .env"
                )
            )

        channel_id = int(
            channel_id
        )

        channel = self.bot.get_channel(
            channel_id
        )

        if channel is None:
            channel = (
                await self.bot.fetch_channel(
                    channel_id
                )
            )

        return channel

    # =========================================================
    # CREATE DRAFTBOARD MESSAGE
    # =========================================================

    async def create_draftboard_message(
        self,
    ):
        draftboard_channel = (
            await self.get_draftboard_channel()
        )

        embed = (
            await self.build_draftboard_embed()
        )

        message = (
            await draftboard_channel.send(
                embed=embed
            )
        )

        save_draftboard_message(
            channel_id=(
                message.channel.id
            ),
            message_id=(
                message.id
            ),
        )

        return message

    # =========================================================
    # CLAIM TEAM AUTOCOMPLETE
    # =========================================================

    async def claim_team_autocomplete(
        self,
        interaction: discord.Interaction,
        current: str,
    ):
        try:
            league = get_league()

            claimed_teams = {
                claim[
                    "espn_team_id"
                ]
                for claim
                in get_all_team_claims()
                if (
                    claim[
                        "discord_user_id"
                    ]
                    != interaction.user.id
                )
            }

            search_text = (
                current
                .lower()
                .strip()
            )

            choices = []

            for team in league.teams:

                if (
                    team.team_id
                    in claimed_teams
                ):
                    continue

                if (
                    search_text
                    and search_text
                    not in team.team_name.lower()
                ):
                    continue

                choices.append(
                    app_commands.Choice(
                        name=(
                            team.team_name
                        ),
                        value=str(
                            team.team_id
                        ),
                    )
                )

            return choices[:25]

        except Exception as error:
            print(
                (
                    "Claim Team Autocomplete Error: "
                    f"{error}"
                )
            )

            return []

    # =========================================================
    # ADMIN CLAIM TEAM AUTOCOMPLETE
    # =========================================================

    async def admin_claim_team_autocomplete(
        self,
        interaction: discord.Interaction,
        current: str,
    ):
        try:
            league = get_league()

            search_text = (
                current
                .lower()
                .strip()
            )

            choices = []

            for team in league.teams:

                if (
                    search_text
                    and search_text
                    not in team.team_name.lower()
                ):
                    continue

                choices.append(
                    app_commands.Choice(
                        name=(
                            team.team_name
                        ),
                        value=str(
                            team.team_id
                        ),
                    )
                )

            return choices[:25]

        except Exception as error:
            print(
                (
                    "Admin Claim Team "
                    "Autocomplete Error: "
                    f"{error}"
                )
            )

            return []

    # =========================================================
    # DRAFT TEAM AUTOCOMPLETE
    # =========================================================

    async def draft_team_autocomplete(
        self,
        interaction: discord.Interaction,
        current: str,
    ):
        try:
            draft_order = (
                get_draft_order()
            )

            search_text = (
                current
                .lower()
                .strip()
            )

            choices = []

            for team in draft_order:

                team_name = (
                    team["team_name"]
                )

                if (
                    search_text
                    and search_text
                    not in team_name.lower()
                ):
                    continue

                choices.append(
                    app_commands.Choice(
                        name=(
                            team_name
                        ),
                        value=str(
                            team[
                                "espn_team_id"
                            ]
                        ),
                    )
                )

            return choices[:25]

        except Exception as error:
            print(
                (
                    "Draft Team Autocomplete "
                    f"Error: {error}"
                )
            )

            return []

    # =========================================================
    # CLAIM TEAM
    # Grants Owner role on success.
    # =========================================================

    @app_commands.command(
        name="claim-team",
        description=(
            "Link your Discord account "
            "to your fantasy team."
        ),
    )
    @app_commands.describe(
        team=(
            "Select your ESPN fantasy team."
        )
    )
    @app_commands.autocomplete(
        team=claim_team_autocomplete
    )
    async def claim_team_command(
        self,
        interaction: discord.Interaction,
        team: str,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        role_was_added = False

        try:
            if not isinstance(
                interaction.user,
                discord.Member,
            ):
                await interaction.followup.send(
                    (
                        "❌ This command can only "
                        "be used inside the server."
                    ),
                    ephemeral=True,
                )
                return

            try:
                team_id = int(
                    team
                )

            except ValueError:
                await interaction.followup.send(
                    (
                        "❌ Please select a team "
                        "from the autocomplete list."
                    ),
                    ephemeral=True,
                )
                return

            league = get_league()

            selected_team = None

            for fantasy_team in league.teams:

                if (
                    fantasy_team.team_id
                    == team_id
                ):
                    selected_team = (
                        fantasy_team
                    )
                    break

            if selected_team is None:
                await interaction.followup.send(
                    (
                        "❌ Could not find that "
                        "ESPN team."
                    ),
                    ephemeral=True,
                )
                return

            existing_team_claim = (
                get_team_claim_by_team(
                    selected_team.team_id
                )
            )

            if (
                existing_team_claim
                and existing_team_claim[
                    "discord_user_id"
                ]
                != interaction.user.id
            ):
                await interaction.followup.send(
                    (
                        f"❌ **{selected_team.team_name}** "
                        "has already been claimed "
                        "by another Discord member."
                    ),
                    ephemeral=True,
                )
                return

            # -------------------------------------------------
            # ADD OWNER ROLE FIRST
            #
            # If database claim fails afterward, we can remove
            # the role again cleanly.
            # -------------------------------------------------

            role_was_added = (
                await self.add_owner_role(
                    interaction.user
                )
            )

            try:
                claim_team(
                    discord_user_id=(
                        interaction.user.id
                    ),
                    espn_team_id=(
                        selected_team.team_id
                    ),
                    team_name=(
                        selected_team.team_name
                    ),
                )

            except Exception:

                if role_was_added:
                    try:
                        await self.remove_owner_role(
                            interaction.user
                        )

                    except Exception as rollback_error:
                        print(
                            (
                                "Owner Role Rollback Error: "
                                f"{rollback_error}"
                            )
                        )

                raise

            await interaction.followup.send(
                (
                    "✅ You are now linked to "
                    f"**{selected_team.team_name}**.\n\n"
                    "You have also been granted "
                    "the **Owner** role."
                ),
                ephemeral=True,
            )

        except discord.Forbidden:
            await interaction.followup.send(
                (
                    "❌ I could not assign the "
                    "**Owner** role.\n\n"
                    "Please make sure the bot has "
                    "**Manage Roles** permission and "
                    "that the bot's role is above "
                    "the Owner role."
                ),
                ephemeral=True,
            )

        except ValueError as error:
            await interaction.followup.send(
                (
                    f"❌ {error}"
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Claim Team Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "claiming that team."
                ),
                ephemeral=True,
            )

    # =========================================================
    # MY TEAM
    # =========================================================

    @app_commands.command(
        name="my-team",
        description=(
            "Show the fantasy team linked "
            "to your Discord account."
        ),
    )
    async def my_team(
        self,
        interaction: discord.Interaction,
    ):
        claim = get_team_claim_by_user(
            interaction.user.id
        )

        if claim is None:
            await interaction.response.send_message(
                (
                    "❌ You have not claimed a "
                    "fantasy team yet.\n"
                    "Use `/claim-team` to link "
                    "your account."
                ),
                ephemeral=True,
            )
            return

        await interaction.response.send_message(
            (
                "🏈 Your fantasy team is "
                f"**{claim['team_name']}**."
            ),
            ephemeral=True,
        )

    # =========================================================
    # UNCLAIM TEAM
    # Removes Owner role.
    # =========================================================

    @app_commands.command(
        name="unclaim-team",
        description=(
            "Remove your fantasy team link."
        ),
    )
    async def unclaim_team(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            if not isinstance(
                interaction.user,
                discord.Member,
            ):
                await interaction.followup.send(
                    (
                        "❌ This command can only "
                        "be used inside the server."
                    ),
                    ephemeral=True,
                )
                return

            claim = get_team_claim_by_user(
                interaction.user.id
            )

            if claim is None:
                await interaction.followup.send(
                    (
                        "❌ You do not currently "
                        "have a claimed team."
                    ),
                    ephemeral=True,
                )
                return

            removed_team_name = (
                claim["team_name"]
            )

            # Remove database claim first.
            remove_team_claim(
                interaction.user.id
            )

            # Then remove server access role.
            await self.remove_owner_role(
                interaction.user
            )

            await interaction.followup.send(
                (
                    "✅ You are no longer linked to "
                    f"**{removed_team_name}**.\n\n"
                    "Your **Owner** role has also "
                    "been removed."
                ),
                ephemeral=True,
            )

        except discord.Forbidden:
            await interaction.followup.send(
                (
                    "⚠️ Your fantasy team claim was "
                    "removed, but I could not remove "
                    "the **Owner** role.\n\n"
                    "A Commissioner or Developer "
                    "should correct the role manually."
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "Unclaim Team Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "removing your team claim."
                ),
                ephemeral=True,
            )

    # =========================================================
    # VIEW CLAIMS
    # =========================================================

    @app_commands.command(
        name="claims",
        description=(
            "Show all Discord-to-team claims."
        ),
    )
    async def claims(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "view team claims."
                    ),
                    ephemeral=True,
                )
                return

            claims = (
                get_all_team_claims()
            )

            if not claims:
                await interaction.followup.send(
                    (
                        "No fantasy teams have "
                        "been claimed yet."
                    ),
                    ephemeral=True,
                )
                return

            embed = discord.Embed(
                title=(
                    "🏈 Fantasy Team Claims"
                ),
            )

            claim_lines = []

            for claim in claims:

                user_id = claim[
                    "discord_user_id"
                ]

                claim_lines.append(
                    (
                        f"**{claim['team_name']}**\n"
                        f"↳ <@{user_id}>"
                    )
                )

            embed.description = (
                "\n\n".join(
                    claim_lines
                )
            )

            await interaction.followup.send(
                embed=embed,
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Claims Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "retrieving team claims."
                ),
                ephemeral=True,
            )

    # =========================================================
    # ADMIN SET CLAIM
    # Grants Owner role to assigned member.
    # Removes Owner from displaced owner.
    # =========================================================

    @app_commands.command(
        name="set-claim",
        description=(
            "Assign a Discord member "
            "to an ESPN fantasy team."
        ),
    )
    @app_commands.describe(
        member=(
            "Discord member to assign."
        ),
        team=(
            "Fantasy team to assign."
        ),
    )
    @app_commands.autocomplete(
        team=admin_claim_team_autocomplete
    )
    async def set_claim(
        self,
        interaction: discord.Interaction,
        member: discord.Member,
        team: str,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        role_was_added = False

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "manage team claims."
                    ),
                    ephemeral=True,
                )
                return

            try:
                team_id = int(
                    team
                )

            except ValueError:
                await interaction.followup.send(
                    (
                        "❌ Please select a team "
                        "from the autocomplete list."
                    ),
                    ephemeral=True,
                )
                return

            league = get_league()

            selected_team = None

            for fantasy_team in league.teams:

                if (
                    fantasy_team.team_id
                    == team_id
                ):
                    selected_team = (
                        fantasy_team
                    )
                    break

            if selected_team is None:
                await interaction.followup.send(
                    (
                        "❌ Could not find that "
                        "ESPN team."
                    ),
                    ephemeral=True,
                )
                return

            existing_team_claim = (
                get_team_claim_by_team(
                    selected_team.team_id
                )
            )

            existing_member_claim = (
                get_team_claim_by_user(
                    member.id
                )
            )

            displaced_member = None

            if (
                existing_team_claim
                and existing_team_claim[
                    "discord_user_id"
                ]
                != member.id
            ):
                displaced_member = (
                    await self.get_guild_member(
                        interaction.guild,
                        existing_team_claim[
                            "discord_user_id"
                        ],
                    )
                )

            # -------------------------------------------------
            # ENSURE NEW MEMBER HAS OWNER ROLE FIRST
            # -------------------------------------------------

            role_was_added = (
                await self.add_owner_role(
                    member
                )
            )

            try:
                if (
                    existing_team_claim
                    and existing_team_claim[
                        "discord_user_id"
                    ]
                    != member.id
                ):
                    remove_team_claim(
                        existing_team_claim[
                            "discord_user_id"
                        ]
                    )

                if (
                    existing_member_claim
                    and existing_member_claim[
                        "espn_team_id"
                    ]
                    != selected_team.team_id
                ):
                    remove_team_claim(
                        member.id
                    )

                claim_team(
                    discord_user_id=(
                        member.id
                    ),
                    espn_team_id=(
                        selected_team.team_id
                    ),
                    team_name=(
                        selected_team.team_name
                    ),
                )

            except Exception:

                if role_was_added:
                    try:
                        await self.remove_owner_role(
                            member
                        )

                    except Exception as rollback_error:
                        print(
                            (
                                "Set Claim Role Rollback Error: "
                                f"{rollback_error}"
                            )
                        )

                raise

            # -------------------------------------------------
            # REMOVE OWNER ROLE FROM DISPLACED MEMBER
            # -------------------------------------------------

            if (
                displaced_member is not None
                and displaced_member.id
                != member.id
            ):
                try:
                    await self.remove_owner_role(
                        displaced_member
                    )

                except discord.Forbidden:
                    print(
                        (
                            "Could not remove Owner role "
                            "from displaced member "
                            f"{displaced_member.id}."
                        )
                    )

            await interaction.followup.send(
                (
                    f"✅ {member.mention} is now "
                    f"linked to "
                    f"**{selected_team.team_name}** "
                    "and has the **Owner** role."
                ),
                ephemeral=True,
            )

        except discord.Forbidden:
            await interaction.followup.send(
                (
                    "❌ I could not manage the "
                    "**Owner** role.\n\n"
                    "Please make sure the bot has "
                    "**Manage Roles** permission and "
                    "that its role is above Owner."
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Set Claim Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "setting that team claim."
                ),
                ephemeral=True,
            )

    # =========================================================
    # ADMIN CLEAR CLAIM
    # Removes Owner role.
    # =========================================================

    @app_commands.command(
        name="clear-claim",
        description=(
            "Remove a Discord member's "
            "fantasy team claim."
        ),
    )
    @app_commands.describe(
        member=(
            "Discord member whose team "
            "claim should be removed."
        )
    )
    async def clear_claim(
        self,
        interaction: discord.Interaction,
        member: discord.Member,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "clear team claims."
                    ),
                    ephemeral=True,
                )
                return

            claim = get_team_claim_by_user(
                member.id
            )

            if claim is None:
                await interaction.followup.send(
                    (
                        f"❌ {member.mention} does "
                        "not currently have a "
                        "claimed fantasy team."
                    ),
                    ephemeral=True,
                )
                return

            removed_team_name = (
                claim["team_name"]
            )

            remove_team_claim(
                member.id
            )

            await self.remove_owner_role(
                member
            )

            await interaction.followup.send(
                (
                    f"✅ Cleared {member.mention}'s "
                    f"claim to "
                    f"**{removed_team_name}**.\n\n"
                    "Their **Owner** role has "
                    "also been removed."
                ),
                ephemeral=True,
            )

        except discord.Forbidden:
            await interaction.followup.send(
                (
                    "⚠️ The database claim was "
                    "removed, but I could not remove "
                    "the member's **Owner** role.\n\n"
                    "Please correct their role "
                    "manually."
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Clear Claim Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "clearing that team claim."
                ),
                ephemeral=True,
            )

    # =========================================================
    # START DRAFT
    # Commissioner / Developer only
    # =========================================================

    @app_commands.command(
        name="start-draft",
        description=(
            "Start a new fantasy football draft."
        ),
    )
    @app_commands.describe(
        rounds=(
            "Number of rounds in the draft."
        )
    )
    async def start_draft(
        self,
        interaction: discord.Interaction,
        rounds: app_commands.Range[
            int,
            1,
            30,
        ],
    ):
        if not self.has_draft_admin_role(
            interaction
        ):
            await interaction.response.send_message(
                (
                    "❌ Only members with the "
                    "**Commissioner** or "
                    "**Developer** role can "
                    "start the draft."
                ),
                ephemeral=True,
            )
            return

        draft_state = (
            get_draft_state()
        )

        if draft_state["active"]:
            await interaction.response.send_message(
                (
                    "❌ A draft is already active.\n"
                    "End or reset the current "
                    "draft before starting another one."
                ),
                ephemeral=True,
            )
            return

        modal = DraftOrderModal(
            draft_cog=self,
            total_rounds=rounds,
        )

        await interaction.response.send_modal(
            modal
        )

    # =========================================================
    # PLAYER AUTOCOMPLETE
    # =========================================================

    async def player_autocomplete(
        self,
        interaction: discord.Interaction,
        current: str,
    ):
        if (
            len(current.strip())
            < 2
        ):
            return []

        try:
            league = get_league()

            drafted_player_ids = (
                get_drafted_player_ids()
            )

            search_text = (
                current
                .lower()
                .strip()
            )

            candidate_ids = []

            for (
                player_name,
                player_id,
            ) in league.player_map.items():

                if not isinstance(
                    player_id,
                    int,
                ):
                    continue

                if (
                    player_id
                    in drafted_player_ids
                ):
                    continue

                if (
                    search_text
                    not in player_name.lower()
                ):
                    continue

                candidate_ids.append(
                    player_id
                )

                if (
                    len(candidate_ids)
                    >= 25
                ):
                    break

            if not candidate_ids:
                return []

            players = league.player_info(
                playerId=candidate_ids
            )

            if players is None:
                return []

            if not isinstance(
                players,
                list,
            ):
                players = [
                    players
                ]

            choices = []

            for player in players:

                if (
                    player.position
                    not in ALLOWED_POSITIONS
                ):
                    continue

                label = (
                    f"{player.name} — "
                    f"{player.position} — "
                    f"{player.proTeam}"
                )

                choices.append(
                    app_commands.Choice(
                        name=label[:100],
                        value=str(
                            player.playerId
                        ),
                    )
                )

                if (
                    len(choices)
                    >= 25
                ):
                    break

            return choices

        except Exception as error:
            print(
                (
                    "Player Autocomplete Error: "
                    f"{error}"
                )
            )

            return []

    # =========================================================
    # BUILD DRAFTBOARD
    # =========================================================

    async def build_draftboard_embed(
        self,
    ):
        state = get_draft_state()

        draft_picks = (
            get_all_draft_picks()
        )

        ownership_rows = (
            get_all_pick_ownership()
        )

        clock_rows = (
            get_all_pick_clocks()
        )

        total_teams = state[
            "total_teams"
        ]

        total_rounds = state[
            "total_rounds"
        ]

        current_pick = state[
            "current_pick"
        ]

        if (
            total_teams is None
            or total_rounds is None
        ):
            raise ValueError(
                (
                    "Draft configuration "
                    "is incomplete."
                )
            )

        ownership_by_pick = {
            row["overall_pick"]: row
            for row in ownership_rows
        }

        picks_by_number = {
            pick["overall_pick"]: pick
            for pick in draft_picks
        }

        clocks_by_number = {
            clock["overall_pick"]: clock
            for clock in clock_rows
        }

        current_round = (
            (
                (
                    current_pick - 1
                )
                // total_teams
            )
            + 1
        )

        if (
            current_round
            > total_rounds
        ):
            current_round = (
                total_rounds
            )

        first_round = max(
            1,
            current_round - 2,
        )

        embed = discord.Embed(
            title=(
                "🏈 Flint Michigan Megabowl "
                "Draft Board"
            ),
            description=(
                f"Showing Rounds "
                f"{first_round}-"
                f"{current_round}\n"
                "🔄 = Traded Pick • "
                "⌛ = Catch-up Pick Owed"
            ),
        )

        for round_number in range(
            first_round,
            current_round + 1,
        ):

            round_lines = []

            for pick_in_round in range(
                1,
                total_teams + 1,
            ):

                overall_pick = (
                    (
                        round_number - 1
                    )
                    * total_teams
                    + pick_in_round
                )

                ownership = (
                    ownership_by_pick.get(
                        overall_pick
                    )
                )

                if ownership is None:
                    continue

                team_name = ownership[
                    "current_team_name"
                ]

                traded = bool(
                    ownership[
                        "traded"
                    ]
                )

                trade_marker = (
                    " 🔄"
                    if traded
                    else ""
                )

                pick = (
                    picks_by_number.get(
                        overall_pick
                    )
                )

                clock = (
                    clocks_by_number.get(
                        overall_pick
                    )
                )

                if pick is not None:
                    status = (
                        f"**{pick['player_name']}** "
                        f"({pick['position']}, "
                        f"{pick['nfl_team']})"
                    )

                elif (
                    clock is not None
                    and clock["status"]
                    == "EXPIRED"
                ):
                    status = (
                        "⌛ **EXPIRED / PICK OWED**"
                    )

                elif (
                    state["active"]
                    and overall_pick
                    == current_pick
                ):
                    status = (
                        "⏰ **ON THE CLOCK**"
                    )

                else:
                    status = "—"

                round_lines.append(
                    (
                        f"`{overall_pick:>3}` "
                        f"**{team_name}**"
                        f"{trade_marker} — "
                        f"{status}"
                    )
                )

            embed.add_field(
                name=(
                    f"Round {round_number}"
                ),
                value="\n".join(
                    round_lines
                ),
                inline=False,
            )

        return embed

    # =========================================================
    # REFRESH DRAFTBOARD
    # =========================================================

    async def refresh_draftboard(
        self,
    ):
        state = get_draft_state()

        channel_id = state[
            "draftboard_channel_id"
        ]

        message_id = state[
            "draftboard_message_id"
        ]

        if (
            not channel_id
            or not message_id
        ):
            return

        try:
            channel = (
                self.bot.get_channel(
                    channel_id
                )
            )

            if channel is None:
                channel = (
                    await self.bot.fetch_channel(
                        channel_id
                    )
                )

            message = (
                await channel.fetch_message(
                    message_id
                )
            )

            embed = (
                await self.build_draftboard_embed()
            )

            await message.edit(
                embed=embed
            )

        except discord.NotFound:
            print(
                (
                    "Draft board message "
                    "no longer exists."
                )
            )

        except Exception as error:
            print(
                (
                    "Draft Board Refresh Error: "
                    f"{error}"
                )
            )

    # =========================================================
    # DRAFTBOARD COMMAND
    # Commissioner / Developer only
    # =========================================================

    @app_commands.command(
        name="draftboard",
        description=(
            "Show or refresh the "
            "current draft board."
        ),
    )
    async def draftboard(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "manually refresh or recreate "
                        "the draft board."
                    ),
                    ephemeral=True,
                )
                return

            state = get_draft_state()

            ownership = (
                get_all_pick_ownership()
            )

            if not ownership:
                await interaction.followup.send(
                    (
                        "❌ No draft has been "
                        "initialized yet."
                    ),
                    ephemeral=True,
                )
                return

            embed = (
                await self.build_draftboard_embed()
            )

            message_id = state[
                "draftboard_message_id"
            ]

            channel_id = state[
                "draftboard_channel_id"
            ]

            if (
                message_id
                and channel_id
            ):
                try:
                    channel = (
                        self.bot.get_channel(
                            channel_id
                        )
                    )

                    if channel is None:
                        channel = (
                            await self.bot.fetch_channel(
                                channel_id
                            )
                        )

                    message = (
                        await channel.fetch_message(
                            message_id
                        )
                    )

                    await message.edit(
                        embed=embed
                    )

                    await interaction.followup.send(
                        (
                            "✅ Draft board "
                            "refreshed."
                        ),
                        ephemeral=True,
                    )

                    return

                except discord.NotFound:
                    pass

            message = (
                await self.create_draftboard_message()
            )

            await interaction.followup.send(
                (
                    "✅ Draft board recreated in "
                    f"{message.channel.mention}."
                ),
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Draft Board Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "updating the draft board."
                ),
                ephemeral=True,
            )

    # =========================================================
    # PICK OWNER
    # =========================================================

    @app_commands.command(
        name="pick-owner",
        description=(
            "Show the current owner "
            "of a draft pick."
        ),
    )
    @app_commands.describe(
        pick=(
            "Overall draft pick number."
        )
    )
    async def pick_owner(
        self,
        interaction: discord.Interaction,
        pick: app_commands.Range[
            int,
            1,
            1000,
        ],
    ):
        await interaction.response.defer()

        try:
            ownership = (
                get_pick_owner(
                    pick
                )
            )

            if ownership is None:
                await interaction.followup.send(
                    (
                        f"❌ Overall pick "
                        f"**{pick}** does not exist."
                    ),
                    ephemeral=True,
                )
                return

            traded = bool(
                ownership[
                    "traded"
                ]
            )

            embed = discord.Embed(
                title=(
                    "🏈 Draft Pick Ownership"
                ),
                description=(
                    f"**Round "
                    f"{ownership['round_number']} "
                    f"• Pick "
                    f"{ownership['pick_in_round']} "
                    f"• Overall {pick}**"
                ),
            )

            embed.add_field(
                name="Current Owner",
                value=(
                    f"**"
                    f"{ownership['current_team_name']}"
                    f"**"
                ),
                inline=False,
            )

            embed.add_field(
                name="Original Owner",
                value=(
                    f"**"
                    f"{ownership['original_team_name']}"
                    f"**"
                ),
                inline=False,
            )

            embed.add_field(
                name="Trade Status",
                value=(
                    "🔄 Traded Pick"
                    if traded
                    else "Original Ownership"
                ),
                inline=False,
            )

            await interaction.followup.send(
                embed=embed
            )

        except Exception as error:
            print(
                f"Pick Owner Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "retrieving pick ownership."
                ),
                ephemeral=True,
            )

    # =========================================================
    # PROPOSE TRADE
    # =========================================================

    @app_commands.command(
        name="propose-trade",
        description=(
            "Offer one of your draft picks "
            "for another team's pick."
        ),
    )
    @app_commands.describe(
        give=(
            "Overall pick you are offering."
        ),
        receive=(
            "Overall pick you want in return."
        ),
    )
    async def propose_trade(
        self,
        interaction: discord.Interaction,
        give: app_commands.Range[
            int,
            1,
            1000,
        ],
        receive: app_commands.Range[
            int,
            1,
            1000,
        ],
    ):
        await interaction.response.defer()

        try:
            state = get_draft_state()

            if not state["active"]:
                await interaction.followup.send(
                    (
                        "❌ There is no active draft."
                    ),
                    ephemeral=True,
                )
                return

            if give == receive:
                await interaction.followup.send(
                    (
                        "❌ You cannot trade a pick "
                        "for itself."
                    ),
                    ephemeral=True,
                )
                return

            proposer_claim = (
                get_team_claim_by_user(
                    interaction.user.id
                )
            )

            if proposer_claim is None:
                await interaction.followup.send(
                    (
                        "❌ You must claim your fantasy "
                        "team before proposing a trade."
                    ),
                    ephemeral=True,
                )
                return

            give_owner = (
                get_pick_owner(
                    give
                )
            )

            receive_owner = (
                get_pick_owner(
                    receive
                )
            )

            if give_owner is None:
                await interaction.followup.send(
                    (
                        f"❌ Overall pick **{give}** "
                        "does not exist."
                    ),
                    ephemeral=True,
                )
                return

            if receive_owner is None:
                await interaction.followup.send(
                    (
                        f"❌ Overall pick **{receive}** "
                        "does not exist."
                    ),
                    ephemeral=True,
                )
                return

            current_pick = state[
                "current_pick"
            ]

            if (
                give < current_pick
                or receive < current_pick
            ):
                await interaction.followup.send(
                    (
                        "❌ Draft picks that have "
                        "already been made cannot "
                        "be traded."
                    ),
                    ephemeral=True,
                )
                return

            if (
                give_owner[
                    "current_espn_team_id"
                ]
                != proposer_claim[
                    "espn_team_id"
                ]
            ):
                await interaction.followup.send(
                    (
                        f"❌ You do not currently own "
                        f"overall pick **{give}**.\n\n"
                        "That pick belongs to "
                        f"**{give_owner['current_team_name']}**."
                    ),
                    ephemeral=True,
                )
                return

            recipient_team_id = (
                receive_owner[
                    "current_espn_team_id"
                ]
            )

            if (
                recipient_team_id
                == proposer_claim[
                    "espn_team_id"
                ]
            ):
                await interaction.followup.send(
                    (
                        "❌ Both picks are already "
                        "owned by your team."
                    ),
                    ephemeral=True,
                )
                return

            recipient_claim = (
                get_team_claim_by_team(
                    recipient_team_id
                )
            )

            if recipient_claim is None:
                await interaction.followup.send(
                    (
                        "❌ The manager of "
                        f"**{receive_owner['current_team_name']}** "
                        "has not claimed their team yet, "
                        "so the bot cannot send them "
                        "a trade proposal."
                    ),
                    ephemeral=True,
                )
                return

            proposal_id = (
                create_trade_proposal(
                    pick_a=give,
                    pick_b=receive,
                    proposer_discord_user_id=(
                        interaction.user.id
                    ),
                    recipient_discord_user_id=(
                        recipient_claim[
                            "discord_user_id"
                        ]
                    ),
                    proposer_team_id=(
                        proposer_claim[
                            "espn_team_id"
                        ]
                    ),
                    proposer_team_name=(
                        proposer_claim[
                            "team_name"
                        ]
                    ),
                    recipient_team_id=(
                        recipient_team_id
                    ),
                    recipient_team_name=(
                        receive_owner[
                            "current_team_name"
                        ]
                    ),
                )
            )

            view = TradeProposalView(
                draft_cog=self,
                proposal_id=(
                    proposal_id
                ),
                recipient_user_id=(
                    recipient_claim[
                        "discord_user_id"
                    ]
                ),
            )

            embed = discord.Embed(
                title=(
                    "🔄 Draft Pick Trade Proposal"
                ),
                description=(
                    f"<@"
                    f"{recipient_claim['discord_user_id']}"
                    f"> — "
                    f"**{proposer_claim['team_name']}** "
                    "has proposed a trade."
                ),
            )

            embed.add_field(
                name=(
                    f"{proposer_claim['team_name']} Gives"
                ),
                value=(
                    f"**Round "
                    f"{give_owner['round_number']} "
                    f"• Pick "
                    f"{give_owner['pick_in_round']} "
                    f"• Overall {give}**"
                ),
                inline=False,
            )

            embed.add_field(
                name=(
                    f"{receive_owner['current_team_name']} Gives"
                ),
                value=(
                    f"**Round "
                    f"{receive_owner['round_number']} "
                    f"• Pick "
                    f"{receive_owner['pick_in_round']} "
                    f"• Overall {receive}**"
                ),
                inline=False,
            )

            embed.set_footer(
                text=(
                    f"Trade Proposal #{proposal_id}"
                )
            )

            proposal_message = (
                await interaction.followup.send(
                    content=(
                        f"<@"
                        f"{recipient_claim['discord_user_id']}"
                        f">"
                    ),
                    embed=embed,
                    view=view,
                    wait=True,
                )
            )

            save_trade_proposal_message(
                proposal_id=(
                    proposal_id
                ),
                discord_channel_id=(
                    proposal_message.channel.id
                ),
                discord_message_id=(
                    proposal_message.id
                ),
            )

        except ValueError as error:
            await interaction.followup.send(
                f"❌ {error}",
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "Propose Trade Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "creating the trade proposal."
                ),
                ephemeral=True,
            )

    # =========================================================
    # MY TRADES
    # =========================================================

    @app_commands.command(
        name="my-trades",
        description=(
            "Show your pending draft-pick "
            "trade proposals."
        ),
    )
    async def my_trades(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            proposals = (
                get_pending_trade_proposals_for_user(
                    interaction.user.id
                )
            )

            if not proposals:
                await interaction.followup.send(
                    (
                        "You do not currently have "
                        "any pending trade proposals."
                    ),
                    ephemeral=True,
                )
                return

            sent = []
            received = []

            for proposal in proposals:

                if (
                    proposal[
                        "proposer_discord_user_id"
                    ]
                    == interaction.user.id
                ):
                    sent.append(
                        proposal
                    )

                if (
                    proposal[
                        "recipient_discord_user_id"
                    ]
                    == interaction.user.id
                ):
                    received.append(
                        proposal
                    )

            embed = discord.Embed(
                title="🔄 My Pending Trades",
                description=(
                    "Use `/cancel-trade` with the "
                    "proposal number to cancel "
                    "one of your sent offers."
                ),
            )

            if sent:

                sent_lines = []

                for proposal in sent:

                    give_owner = get_pick_owner(
                        proposal[
                            "pick_a"
                        ]
                    )

                    receive_owner = get_pick_owner(
                        proposal[
                            "pick_b"
                        ]
                    )

                    if (
                        give_owner is not None
                        and receive_owner is not None
                    ):
                        sent_lines.append(
                            (
                                f"**Trade #{proposal['id']}** "
                                f"— vs "
                                f"**{proposal['recipient_team_name']}**\n"
                                f"You Give: Round "
                                f"{give_owner['round_number']} • "
                                f"Pick "
                                f"{give_owner['pick_in_round']} "
                                f"(Overall "
                                f"{proposal['pick_a']})\n"
                                f"You Receive: Round "
                                f"{receive_owner['round_number']} • "
                                f"Pick "
                                f"{receive_owner['pick_in_round']} "
                                f"(Overall "
                                f"{proposal['pick_b']})"
                            )
                        )

                    else:
                        sent_lines.append(
                            (
                                f"**Trade #{proposal['id']}** "
                                f"— vs "
                                f"**{proposal['recipient_team_name']}**\n"
                                f"You Give: Overall "
                                f"{proposal['pick_a']}\n"
                                f"You Receive: Overall "
                                f"{proposal['pick_b']}"
                            )
                        )

                embed.add_field(
                    name="📤 Sent",
                    value="\n\n".join(
                        sent_lines
                    ),
                    inline=False,
                )

            if received:

                received_lines = []

                for proposal in received:

                    offered_owner = get_pick_owner(
                        proposal[
                            "pick_a"
                        ]
                    )

                    your_pick_owner = get_pick_owner(
                        proposal[
                            "pick_b"
                        ]
                    )

                    if (
                        offered_owner is not None
                        and your_pick_owner is not None
                    ):
                        received_lines.append(
                            (
                                f"**Trade #{proposal['id']}** "
                                f"— from "
                                f"**{proposal['proposer_team_name']}**\n"
                                f"They Give: Round "
                                f"{offered_owner['round_number']} • "
                                f"Pick "
                                f"{offered_owner['pick_in_round']} "
                                f"(Overall "
                                f"{proposal['pick_a']})\n"
                                f"You Give: Round "
                                f"{your_pick_owner['round_number']} • "
                                f"Pick "
                                f"{your_pick_owner['pick_in_round']} "
                                f"(Overall "
                                f"{proposal['pick_b']})"
                            )
                        )

                    else:
                        received_lines.append(
                            (
                                f"**Trade #{proposal['id']}** "
                                f"— from "
                                f"**{proposal['proposer_team_name']}**\n"
                                f"They Give: Overall "
                                f"{proposal['pick_a']}\n"
                                f"You Give: Overall "
                                f"{proposal['pick_b']}"
                            )
                        )

                embed.add_field(
                    name="📥 Received",
                    value="\n\n".join(
                        received_lines
                    ),
                    inline=False,
                )

            await interaction.followup.send(
                embed=embed,
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "My Trades Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "retrieving your trade proposals."
                ),
                ephemeral=True,
            )

    # =========================================================
    # UPDATE CANCELLED TRADE MESSAGE
    # =========================================================

    async def update_cancelled_trade_message(
        self,
        proposal,
    ):
        channel_id = proposal[
            "discord_channel_id"
        ]

        message_id = proposal[
            "discord_message_id"
        ]

        if (
            not channel_id
            or not message_id
        ):
            return False

        try:
            channel = self.bot.get_channel(
                channel_id
            )

            if channel is None:
                channel = (
                    await self.bot.fetch_channel(
                        channel_id
                    )
                )

            message = (
                await channel.fetch_message(
                    message_id
                )
            )

            give_owner = get_pick_owner(
                proposal["pick_a"]
            )

            receive_owner = get_pick_owner(
                proposal["pick_b"]
            )

            embed = discord.Embed(
                title=(
                    "🚫 Draft Pick Trade Cancelled"
                ),
                description=(
                    f"**{proposal['proposer_team_name']}** "
                    "cancelled this trade proposal."
                ),
            )

            if give_owner is not None:
                embed.add_field(
                    name=(
                        f"{proposal['proposer_team_name']} "
                        "Would Have Given"
                    ),
                    value=(
                        f"**Round "
                        f"{give_owner['round_number']} "
                        f"• Pick "
                        f"{give_owner['pick_in_round']} "
                        f"• Overall "
                        f"{proposal['pick_a']}**"
                    ),
                    inline=False,
                )

            else:
                embed.add_field(
                    name=(
                        f"{proposal['proposer_team_name']} "
                        "Would Have Given"
                    ),
                    value=(
                        f"**Overall "
                        f"{proposal['pick_a']}**"
                    ),
                    inline=False,
                )

            if receive_owner is not None:
                embed.add_field(
                    name=(
                        f"{proposal['recipient_team_name']} "
                        "Would Have Given"
                    ),
                    value=(
                        f"**Round "
                        f"{receive_owner['round_number']} "
                        f"• Pick "
                        f"{receive_owner['pick_in_round']} "
                        f"• Overall "
                        f"{proposal['pick_b']}**"
                    ),
                    inline=False,
                )

            else:
                embed.add_field(
                    name=(
                        f"{proposal['recipient_team_name']} "
                        "Would Have Given"
                    ),
                    value=(
                        f"**Overall "
                        f"{proposal['pick_b']}**"
                    ),
                    inline=False,
                )

            embed.add_field(
                name="Status",
                value="🚫 **CANCELLED**",
                inline=False,
            )

            embed.set_footer(
                text=(
                    f"Trade Proposal "
                    f"#{proposal['id']}"
                )
            )

            disabled_view = TradeProposalView(
                draft_cog=self,
                proposal_id=(
                    proposal["id"]
                ),
                recipient_user_id=(
                    proposal[
                        "recipient_discord_user_id"
                    ]
                ),
            )

            disabled_view.disable_all_buttons()

            await message.edit(
                content=None,
                embed=embed,
                view=disabled_view,
            )

            return True

        except discord.NotFound:
            print(
                (
                    "Cancelled trade proposal "
                    f"message #{proposal['id']} "
                    "no longer exists."
                )
            )

            return False

        except Exception as error:
            print(
                (
                    "Cancelled Trade Message "
                    f"Update Error: {error}"
                )
            )

            return False

    # =========================================================
    # CANCEL TRADE
    # =========================================================

    @app_commands.command(
        name="cancel-trade",
        description=(
            "Cancel one of your pending "
            "draft-pick trade proposals."
        ),
    )
    @app_commands.describe(
        proposal=(
            "Trade proposal number."
        )
    )
    async def cancel_trade(
        self,
        interaction: discord.Interaction,
        proposal: app_commands.Range[
            int,
            1,
            1000000,
        ],
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            admin_override = (
                self.has_draft_admin_role(
                    interaction
                )
            )

            cancelled = (
                cancel_trade_proposal(
                    proposal_id=(
                        proposal
                    ),
                    cancelling_discord_user_id=(
                        interaction.user.id
                    ),
                    admin_override=(
                        admin_override
                    ),
                )
            )

            message_updated = (
                await self.update_cancelled_trade_message(
                    cancelled
                )
            )

            message_note = (
                "The original trade proposal "
                "message was updated."
                if message_updated
                else (
                    "The trade was cancelled, but "
                    "the original Discord message "
                    "could not be updated."
                )
            )

            await interaction.followup.send(
                (
                    f"✅ Trade Proposal "
                    f"**#{proposal}** was cancelled.\n\n"
                    f"**{cancelled['proposer_team_name']}** "
                    f"offered Overall "
                    f"**{cancelled['pick_a']}** for "
                    f"**{cancelled['recipient_team_name']}**'s "
                    f"Overall **{cancelled['pick_b']}**.\n\n"
                    f"{message_note}"
                ),
                ephemeral=True,
            )

        except ValueError as error:
            await interaction.followup.send(
                f"❌ {error}",
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "Cancel Trade Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "cancelling the trade proposal."
                ),
                ephemeral=True,
            )

    # =========================================================
    # ADMIN TRADE PICKS
    # =========================================================

    @app_commands.command(
        name="trade-picks",
        description=(
            "Admin override: swap two future draft picks."
        ),
    )
    @app_commands.describe(
        pick_a=(
            "First overall pick in the trade."
        ),
        pick_b=(
            "Second overall pick in the trade."
        ),
    )
    async def trade_picks(
        self,
        interaction: discord.Interaction,
        pick_a: app_commands.Range[
            int,
            1,
            1000,
        ],
        pick_b: app_commands.Range[
            int,
            1,
            1000,
        ],
    ):
        await interaction.response.defer()

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "record draft-pick trades directly."
                    ),
                    ephemeral=True,
                )
                return

            result = trade_draft_picks(
                pick_a=pick_a,
                pick_b=pick_b,
                discord_user_id=(
                    interaction.user.id
                ),
            )

            invalidate_pending_trade_proposals_for_picks(
                pick_a=pick_a,
                pick_b=pick_b,
            )

            first = result[
                "pick_a"
            ]

            second = result[
                "pick_b"
            ]

            embed = discord.Embed(
                title=(
                    "🔄 Draft Pick Trade"
                ),
                description=(
                    "Pick ownership has been swapped."
                ),
            )

            embed.add_field(
                name=(
                    f"Overall {first['overall_pick']} "
                    f"— Round "
                    f"{first['round_number']} "
                    f"Pick "
                    f"{first['pick_in_round']}"
                ),
                value=(
                    f"~~{first['previous_team_name']}~~\n"
                    f"➡️ **{first['new_team_name']}**"
                ),
                inline=False,
            )

            embed.add_field(
                name=(
                    f"Overall {second['overall_pick']} "
                    f"— Round "
                    f"{second['round_number']} "
                    f"Pick "
                    f"{second['pick_in_round']}"
                ),
                value=(
                    f"~~{second['previous_team_name']}~~\n"
                    f"➡️ **{second['new_team_name']}**"
                ),
                inline=False,
            )

            await interaction.followup.send(
                embed=embed
            )

            await self.refresh_draftboard()

        except ValueError as error:
            await interaction.followup.send(
                f"❌ {error}",
                ephemeral=True,
            )

        except Exception as error:
            print(
                f"Trade Picks Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "recording the pick trade."
                ),
                ephemeral=True,
            )

    # =========================================================
    # UNDO PICK TRADE
    # =========================================================

    @app_commands.command(
        name="undo-pick-trade",
        description=(
            "Undo the latest trade involving a draft pick."
        ),
    )
    @app_commands.describe(
        pick=(
            "Overall pick involved in the trade."
        )
    )
    async def undo_pick_trade(
        self,
        interaction: discord.Interaction,
        pick: app_commands.Range[
            int,
            1,
            1000,
        ],
    ):
        await interaction.response.defer()

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "undo pick trades."
                    ),
                    ephemeral=True,
                )
                return

            result = (
                undo_last_pick_trade(
                    pick
                )
            )

            if result is None:
                await interaction.followup.send(
                    (
                        f"❌ Overall pick **{pick}** "
                        "has no trade to undo."
                    ),
                    ephemeral=True,
                )
                return

            first = result[
                "pick_a"
            ]

            second = result[
                "pick_b"
            ]

            embed = discord.Embed(
                title=(
                    "↩️ Draft Pick Trade Undone"
                ),
                description=(
                    "Both picks have been restored "
                    "to their pre-trade owners."
                ),
            )

            embed.add_field(
                name=(
                    f"Overall "
                    f"{first['overall_pick']}"
                ),
                value=(
                    f"**{first['team_name']}**"
                ),
                inline=False,
            )

            embed.add_field(
                name=(
                    f"Overall "
                    f"{second['overall_pick']}"
                ),
                value=(
                    f"**{second['team_name']}**"
                ),
                inline=False,
            )

            await interaction.followup.send(
                embed=embed
            )

            await self.refresh_draftboard()

        except ValueError as error:
            await interaction.followup.send(
                f"❌ {error}",
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "Undo Pick Trade Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "undoing the pick trade."
                ),
                ephemeral=True,
            )

    # =========================================================
    # TEAM ROSTER
    # =========================================================

    @app_commands.command(
        name="team-roster",
        description=(
            "Show the players drafted "
            "by a fantasy team."
        ),
    )
    @app_commands.describe(
        team=(
            "Optional: select a fantasy team. "
            "Defaults to your claimed team."
        )
    )
    @app_commands.autocomplete(
        team=draft_team_autocomplete
    )
    async def team_roster(
        self,
        interaction: discord.Interaction,
        team: str | None = None,
    ):
        await interaction.response.defer()

        try:
            if team is None:

                claim = (
                    get_team_claim_by_user(
                        interaction.user.id
                    )
                )

                if claim is None:
                    await interaction.followup.send(
                        (
                            "❌ You have not claimed a "
                            "fantasy team yet.\n"
                            "Use `/claim-team`, or specify "
                            "a team with `/team-roster`."
                        ),
                        ephemeral=True,
                    )
                    return

                team_id = claim[
                    "espn_team_id"
                ]

            else:

                try:
                    team_id = int(
                        team
                    )

                except ValueError:
                    await interaction.followup.send(
                        (
                            "❌ Please select a team "
                            "from the autocomplete list."
                        ),
                        ephemeral=True,
                    )
                    return

            draft_order = (
                get_draft_order()
            )

            selected_team = None

            for fantasy_team in draft_order:

                if (
                    fantasy_team[
                        "espn_team_id"
                    ]
                    == team_id
                ):
                    selected_team = (
                        fantasy_team
                    )
                    break

            if selected_team is None:
                await interaction.followup.send(
                    (
                        "❌ Could not find that "
                        "fantasy team."
                    ),
                    ephemeral=True,
                )
                return

            draft_picks = (
                get_all_draft_picks()
            )

            team_picks = [
                pick
                for pick in draft_picks
                if (
                    pick[
                        "espn_team_id"
                    ]
                    == team_id
                )
            ]

            embed = discord.Embed(
                title=(
                    f"🏈 "
                    f"{selected_team['team_name']} "
                    "Drafted Roster"
                ),
            )

            if not team_picks:
                embed.description = (
                    "No players have been drafted yet."
                )

            else:
                roster_lines = []

                for pick in team_picks:

                    roster_lines.append(
                        (
                            f"**{pick['player_name']}** "
                            f"({pick['position']} • "
                            f"{pick['nfl_team']})\n"
                            f"↳ Round "
                            f"{pick['round_number']} • "
                            f"Overall "
                            f"{pick['overall_pick']}"
                        )
                    )

                embed.description = (
                    "\n\n".join(
                        roster_lines
                    )
                )

            await interaction.followup.send(
                embed=embed
            )

        except Exception as error:
            print(
                (
                    "Team Roster Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "retrieving that draft roster."
                ),
                ephemeral=True,
            )

    # =========================================================
    # AVAILABLE PLAYERS
    # =========================================================

    @app_commands.command(
        name="available",
        description=(
            "Search for top undrafted players."
        ),
    )
    @app_commands.describe(
        search=(
            "Optional player name search."
        ),
        position=(
            "Optional position filter."
        ),
    )
    @app_commands.choices(
        position=[
            app_commands.Choice(
                name="QB",
                value="QB",
            ),
            app_commands.Choice(
                name="RB",
                value="RB",
            ),
            app_commands.Choice(
                name="WR",
                value="WR",
            ),
            app_commands.Choice(
                name="TE",
                value="TE",
            ),
            app_commands.Choice(
                name="K",
                value="K",
            ),
            app_commands.Choice(
                name="D/ST",
                value="D/ST",
            ),
            app_commands.Choice(
                name="DB",
                value="DB",
            ),
            app_commands.Choice(
                name="LB",
                value="LB",
            ),
        ]
    )
    async def available(
        self,
        interaction: discord.Interaction,
        search: str | None = None,
        position: (
            app_commands.Choice[str]
            | None
        ) = None,
    ):
        await interaction.response.defer()

        try:
            if (
                not search
                and position is None
            ):
                await interaction.followup.send(
                    (
                        "❌ Please provide either "
                        "a player name or a "
                        "position filter."
                    ),
                    ephemeral=True,
                )
                return

            league = get_league()

            drafted_player_ids = (
                get_drafted_player_ids()
            )

            search_text = (
                search.lower().strip()
                if search
                else None
            )

            position_filter = (
                position.value
                if position
                else None
            )

            ranked_players = (
                self.get_ranked_player_pool(
                    league=league,
                    drafted_player_ids=(
                        drafted_player_ids
                    ),
                    position_filter=(
                        position_filter
                    ),
                    search_text=(
                        search_text
                    ),
                )
            )

            if not ranked_players:
                await interaction.followup.send(
                    (
                        "No matching undrafted "
                        "players were found."
                    )
                )
                return

            top_players = (
                ranked_players[:15]
            )

            top_player_ids = [
                player[
                    "player_id"
                ]
                for player
                in top_players
            ]

            details = (
                league.player_info(
                    playerId=(
                        top_player_ids
                    )
                )
            )

            if details is None:
                details = []

            elif not isinstance(
                details,
                list,
            ):
                details = [
                    details
                ]

            details_by_id = {
                player.playerId: player
                for player in details
            }

            if (
                position_filter
                and not search
            ):
                embed = discord.Embed(
                    title=(
                        f"🏈 Top 15 Available "
                        f"{position_filter}s"
                    ),
                    description=(
                        "Sorted by ESPN consensus "
                        "position rank."
                    ),
                )

            else:
                embed = discord.Embed(
                    title=(
                        "🔎 Available Player Search"
                    ),
                )

                description_parts = []

                if search:
                    description_parts.append(
                        (
                            f'Name contains '
                            f'**"{search}"**'
                        )
                    )

                if position_filter:
                    description_parts.append(
                        (
                            "Position: "
                            f"**{position_filter}**"
                        )
                    )

                embed.description = (
                    " • ".join(
                        description_parts
                    )
                )

            player_lines = []

            for ranked_player in top_players:

                player_id = (
                    ranked_player[
                        "player_id"
                    ]
                )

                detail = (
                    details_by_id.get(
                        player_id
                    )
                )

                if detail is not None:

                    player_name = (
                        detail.name
                    )

                    player_position = (
                        detail.position
                    )

                    nfl_team = (
                        detail.proTeam
                    )

                else:

                    player_name = (
                        ranked_player[
                            "name"
                        ]
                    )

                    player_position = (
                        position_filter
                        if position_filter
                        else "?"
                    )

                    nfl_team = "?"

                rank_text = (
                    self.format_position_rank(
                        player_position,
                        ranked_player[
                            "rank"
                        ],
                    )
                )

                player_lines.append(
                    (
                        f"**{player_name}** — "
                        f"{player_position} • "
                        f"{nfl_team}\n"
                        "↳ ESPN Position Rank: "
                        f"**{rank_text}**"
                    )
                )

            embed.add_field(
                name=(
                    f"Results "
                    f"({len(top_players)})"
                ),
                value="\n\n".join(
                    player_lines
                ),
                inline=False,
            )

            if (
                len(ranked_players)
                > 15
            ):
                embed.set_footer(
                    text=(
                        "Showing 15 of "
                        f"{len(ranked_players)} "
                        "matching undrafted players."
                    )
                )

            await interaction.followup.send(
                embed=embed
            )

        except Exception as error:
            print(
                (
                    "Available Players Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "searching available players."
                ),
                ephemeral=True,
            )

    # =========================================================
    # MAKE PICK
    # =========================================================

    @app_commands.command(
        name="pick",
        description=(
            "Draft a player for the team "
            "currently on the clock."
        ),
    )
    @app_commands.describe(
        player=(
            "Search for the player being drafted."
        )
    )
    @app_commands.autocomplete(
        player=player_autocomplete
    )
    async def pick(
        self,
        interaction: discord.Interaction,
        player: str,
    ):
        await interaction.response.defer()

        try:
            state = get_draft_state()

            if not state["active"]:
                await interaction.followup.send(
                    (
                        "❌ There is no active draft."
                    ),
                    ephemeral=True,
                )
                return

            current = (
                get_current_draft_team()
            )

            if current is None:
                await interaction.followup.send(
                    (
                        "❌ There is no active "
                        "draft pick."
                    ),
                    ephemeral=True,
                )
                return

            target_overall_pick = (
                current["overall_pick"]
            )
            catch_up_pick = None

            if not self.has_draft_admin_role(
                interaction
            ):

                claim = (
                    get_team_claim_by_user(
                        interaction.user.id
                    )
                )

                if claim is None:
                    await interaction.followup.send(
                        (
                            "❌ You have not claimed "
                            "a fantasy team yet.\n"
                            "Use `/claim-team` before "
                            "making a pick."
                        ),
                        ephemeral=True,
                    )
                    return

                current_team_id = (
                    current["team"][
                        "espn_team_id"
                    ]
                )

                if (
                    claim["espn_team_id"]
                    != current_team_id
                ):
                    catch_up_pick = (
                        get_oldest_expired_pick_for_team(
                            claim["espn_team_id"]
                        )
                    )

                    if catch_up_pick is None:
                        await interaction.followup.send(
                            (
                                "❌ It is not your turn "
                                "to pick.\n\n"
                                f"**{current['team']['team_name']}** "
                                "is currently on the clock."
                            ),
                            ephemeral=True,
                        )
                        return

                    target_overall_pick = (
                        catch_up_pick[
                            "overall_pick"
                        ]
                    )

            try:
                player_id = int(
                    player
                )

            except ValueError:
                await interaction.followup.send(
                    (
                        "❌ Please select a player "
                        "from the autocomplete list."
                    ),
                    ephemeral=True,
                )
                return

            drafted_player_ids = (
                get_drafted_player_ids()
            )

            if (
                player_id
                in drafted_player_ids
            ):
                await interaction.followup.send(
                    (
                        "❌ That player has "
                        "already been drafted."
                    ),
                    ephemeral=True,
                )
                return

            league = get_league()

            selected_player = (
                league.player_info(
                    playerId=player_id
                )
            )

            if selected_player is None:
                await interaction.followup.send(
                    (
                        "❌ Could not find that "
                        "ESPN player."
                    ),
                    ephemeral=True,
                )
                return

            if (
                selected_player.position
                not in ALLOWED_POSITIONS
            ):
                await interaction.followup.send(
                    (
                        f"❌ {selected_player.name} "
                        "has position "
                        f"`{selected_player.position}`, "
                        "which is not draftable "
                        "in this league."
                    ),
                    ephemeral=True,
                )
                return

            saved_pick = (
                save_draft_pick(
                    espn_player_id=(
                        selected_player.playerId
                    ),
                    player_name=(
                        selected_player.name
                    ),
                    position=(
                        selected_player.position
                    ),
                    nfl_team=(
                        selected_player.proTeam
                    ),
                    discord_user_id=(
                        interaction.user.id
                    ),
                    target_overall_pick=(
                        target_overall_pick
                    ),
                )
            )

            team = (
                saved_pick[
                    "team"
                ]
            )

            embed = discord.Embed(
                title=(
                    "⌛ Catch-Up Draft Pick"
                    if saved_pick.get("catch_up")
                    else "🏈 Draft Pick"
                ),
                description=(
                    f"**Round "
                    f"{saved_pick['round_number']} "
                    f"• Pick "
                    f"{saved_pick['pick_in_round']} "
                    f"• Overall "
                    f"{saved_pick['overall_pick']}**"
                ),
            )

            pick_marker = (
                " 🔄"
                if saved_pick.get(
                    "traded"
                )
                else ""
            )

            embed.add_field(
                name=(
                    f"{team['team_name']}"
                    f"{pick_marker}"
                ),
                value=(
                    f"**{selected_player.name}**\n"
                    f"{selected_player.position} • "
                    f"{selected_player.proTeam}"
                ),
                inline=False,
            )

            if saved_pick.get("catch_up"):
                embed.add_field(
                    name="⌛ Catch-Up Pick",
                    value=(
                        "This fills the team's oldest "
                        "expired draft pick. The active "
                        "draft clock was not changed."
                    ),
                    inline=False,
                )

            new_state = (
                get_draft_state()
            )

            if not new_state["active"]:

                embed.add_field(
                    name="🏆 Draft Complete",
                    value=(
                        "The final pick has been made. "
                        "The draft is now complete."
                    ),
                    inline=False,
                )

            else:

                next_pick = (
                    get_current_draft_team()
                )

                next_marker = (
                    " 🔄"
                    if next_pick[
                        "traded"
                    ]
                    else ""
                )

                embed.add_field(
                    name="⏰ On the Clock",
                    value=(
                        f"**"
                        f"{next_pick['team']['team_name']}"
                        f"{next_marker}"
                        f"**\n"
                        f"Round "
                        f"{next_pick['round_number']} "
                        f"• Pick "
                        f"{next_pick['pick_in_round']} "
                        f"• Overall "
                        f"{next_pick['overall_pick']}"
                    ),
                    inline=False,
                )

            await interaction.followup.send(
                embed=embed
            )

            await self.refresh_draftboard()

            if not saved_pick.get("catch_up"):
                await self.process_draft_clock()

        except Exception as error:
            print(
                f"Pick Error: {error}"
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "recording the pick."
                ),
                ephemeral=True,
            )

    # =========================================================
    # UNDO PICK
    # =========================================================

    @app_commands.command(
        name="undo-pick",
        description=(
            "Undo the most recent draft pick."
        ),
    )
    async def undo_pick(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer()

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "undo draft picks."
                    ),
                    ephemeral=True,
                )
                return

            undone_pick = (
                undo_last_draft_pick()
            )

            if undone_pick is None:
                await interaction.followup.send(
                    (
                        "❌ There are no draft "
                        "picks to undo."
                    ),
                    ephemeral=True,
                )
                return

            await self.refresh_draftboard()

            embed = discord.Embed(
                title="↩️ Draft Pick Undone",
                description=(
                    f"**{undone_pick['player_name']}** "
                    f"({undone_pick['position']} • "
                    f"{undone_pick['nfl_team']})"
                ),
            )

            embed.add_field(
                name="Team",
                value=(
                    undone_pick[
                        "team_name"
                    ]
                ),
                inline=False,
            )

            embed.add_field(
                name="Pick",
                value=(
                    f"Round "
                    f"{undone_pick['round_number']} "
                    f"• Pick "
                    f"{undone_pick['pick_in_round']} "
                    f"• Overall "
                    f"{undone_pick['overall_pick']}"
                ),
                inline=False,
            )

            current = (
                get_current_draft_team()
            )

            if current is not None:

                embed.add_field(
                    name="⏰ Back On the Clock",
                    value=(
                        f"**"
                        f"{current['team']['team_name']}"
                        f"**\n"
                        f"Overall Pick "
                        f"{current['overall_pick']}"
                    ),
                    inline=False,
                )

            await interaction.followup.send(
                embed=embed
            )

        except Exception as error:
            print(
                (
                    "Undo Pick Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "undoing the pick."
                ),
                ephemeral=True,
            )

    # =========================================================
    # DRAFT STATUS
    # =========================================================

    @app_commands.command(
        name="draft-status",
        description=(
            "Show the current status "
            "of the fantasy draft."
        ),
    )
    async def draft_status(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer()

        try:
            state = get_draft_state()

            ownership = (
                get_all_pick_ownership()
            )

            draft_picks = (
                get_all_draft_picks()
            )

            if not ownership:
                await interaction.followup.send(
                    (
                        "❌ No draft has been "
                        "initialized yet."
                    ),
                    ephemeral=True,
                )
                return

            total_teams = state[
                "total_teams"
            ]

            total_rounds = state[
                "total_rounds"
            ]

            if (
                total_teams is None
                or total_rounds is None
            ):
                await interaction.followup.send(
                    (
                        "❌ Draft configuration "
                        "is incomplete."
                    ),
                    ephemeral=True,
                )
                return

            total_picks = (
                total_teams
                * total_rounds
            )

            picks_completed = len(
                draft_picks
            )

            picks_remaining = max(
                0,
                (
                    total_picks
                    - picks_completed
                ),
            )

            traded_pick_count = sum(
                1
                for row in ownership
                if row["traded"]
            )

            catch_up_count = (
                get_outstanding_expired_pick_count()
            )

            embed = discord.Embed(
                title=(
                    "🏈 Flint Michigan Megabowl "
                    "Draft Status"
                ),
            )

            embed.add_field(
                name="Format",
                value=(
                    "Snake Draft\n"
                    f"{total_teams} Teams • "
                    f"{total_rounds} Rounds"
                ),
                inline=True,
            )

            embed.add_field(
                name="Progress",
                value=(
                    f"{picks_completed} / "
                    f"{total_picks} Picks\n"
                    f"{picks_remaining} Remaining"
                ),
                inline=True,
            )

            embed.add_field(
                name="Traded Picks",
                value=str(
                    traded_pick_count
                ),
                inline=True,
            )

            embed.add_field(
                name="Catch-Up Picks Owed",
                value=str(
                    catch_up_count
                ),
                inline=True,
            )

            if state["active"]:

                current = (
                    get_current_draft_team()
                )

                embed.add_field(
                    name="Status",
                    value=(
                        "🟢 Draft Active"
                    ),
                    inline=True,
                )

                if current is not None:

                    trade_marker = (
                        " 🔄"
                        if current[
                            "traded"
                        ]
                        else ""
                    )

                    embed.add_field(
                        name="⏰ On the Clock",
                        value=(
                            f"**"
                            f"{current['team']['team_name']}"
                            f"{trade_marker}"
                            f"**\n"
                            f"Round "
                            f"{current['round_number']} "
                            f"• Pick "
                            f"{current['pick_in_round']} "
                            f"• Overall "
                            f"{current['overall_pick']}"
                        ),
                        inline=False,
                    )

                    current_clock = (
                        get_current_pick_clock()
                    )
                    deadline = self.format_clock_deadline(
                        current_clock
                    )

                    if deadline:
                        embed.add_field(
                            name="Draft Clock",
                            value=(
                                f"Expires {deadline}"
                            ),
                            inline=False,
                        )

            else:

                embed.add_field(
                    name="Status",
                    value=(
                        "🏁 Draft Complete"
                    ),
                    inline=True,
                )

            if draft_picks:

                last_pick = (
                    draft_picks[-1]
                )

                embed.add_field(
                    name="Most Recent Pick",
                    value=(
                        f"**{last_pick['player_name']}** "
                        f"({last_pick['position']} • "
                        f"{last_pick['nfl_team']})\n"
                        f"{last_pick['team_name']} • "
                        f"Overall "
                        f"{last_pick['overall_pick']}"
                    ),
                    inline=False,
                )

            await interaction.followup.send(
                embed=embed
            )

        except Exception as error:
            print(
                (
                    "Draft Status Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "retrieving the draft status."
                ),
                ephemeral=True,
            )

    # =========================================================
    # EXPORT HELPERS
    # =========================================================

    def format_export_worksheet(
        self,
        worksheet,
    ):
        for cell in worksheet[1]:
            cell.font = Font(
                bold=True
            )

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for column in worksheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                value = (
                    ""
                    if cell.value is None
                    else str(
                        cell.value
                    )
                )

                max_length = max(
                    max_length,
                    len(value),
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                45,
            )

    # =========================================================
    # EXPORT DRAFT
    # =========================================================

    @app_commands.command(
        name="export-draft",
        description=(
            "Export the current draft results."
        ),
    )
    async def export_draft(
        self,
        interaction: discord.Interaction,
    ):
        await interaction.response.defer(
            ephemeral=True
        )

        try:
            if not self.has_draft_admin_role(
                interaction
            ):
                await interaction.followup.send(
                    (
                        "❌ Only members with the "
                        "**Commissioner** or "
                        "**Developer** role can "
                        "export the draft."
                    ),
                    ephemeral=True,
                )
                return

            draft_picks = (
                get_all_draft_picks()
            )

            if not draft_picks:
                await interaction.followup.send(
                    (
                        "❌ There are no draft "
                        "picks to export."
                    ),
                    ephemeral=True,
                )
                return

            ownership_rows = (
                get_all_pick_ownership()
            )

            trade_rows = (
                get_all_pick_trades()
            )

            draft_order = (
                get_draft_order()
            )

            export_directory = Path(
                "exports"
            )

            export_directory.mkdir(
                exist_ok=True
            )

            csv_path = (
                export_directory
                / "megabowl_draft_results.csv"
            )

            excel_path = (
                export_directory
                / "megabowl_draft_results.xlsx"
            )

            draft_headers = [
                "Overall Pick",
                "Round",
                "Pick in Round",
                "ESPN Team ID",
                "Fantasy Team",
                "ESPN Player ID",
                "Player Name",
                "Position",
                "NFL Team",
                "Picked At",
            ]

            # -------------------------------------------------
            # CSV
            # -------------------------------------------------

            with open(
                csv_path,
                "w",
                newline="",
                encoding="utf-8",
            ) as csv_file:

                writer = csv.writer(
                    csv_file
                )

                writer.writerow(
                    draft_headers
                )

                for pick in draft_picks:

                    writer.writerow(
                        [
                            pick[
                                "overall_pick"
                            ],
                            pick[
                                "round_number"
                            ],
                            pick[
                                "pick_in_round"
                            ],
                            pick[
                                "espn_team_id"
                            ],
                            pick[
                                "team_name"
                            ],
                            pick[
                                "espn_player_id"
                            ],
                            pick[
                                "player_name"
                            ],
                            pick[
                                "position"
                            ],
                            pick[
                                "nfl_team"
                            ],
                            pick[
                                "picked_at"
                            ],
                        ]
                    )

            # -------------------------------------------------
            # WORKBOOK
            # -------------------------------------------------

            workbook = Workbook()

            # =================================================
            # DRAFT RESULTS
            # =================================================

            draft_sheet = (
                workbook.active
            )

            draft_sheet.title = (
                "Draft Results"
            )

            draft_sheet.append(
                draft_headers
            )

            for pick in draft_picks:

                draft_sheet.append(
                    [
                        pick[
                            "overall_pick"
                        ],
                        pick[
                            "round_number"
                        ],
                        pick[
                            "pick_in_round"
                        ],
                        pick[
                            "espn_team_id"
                        ],
                        pick[
                            "team_name"
                        ],
                        pick[
                            "espn_player_id"
                        ],
                        pick[
                            "player_name"
                        ],
                        pick[
                            "position"
                        ],
                        pick[
                            "nfl_team"
                        ],
                        pick[
                            "picked_at"
                        ],
                    ]
                )

            self.format_export_worksheet(
                draft_sheet
            )

            # =================================================
            # BY TEAM
            # =================================================

            team_sheet = (
                workbook.create_sheet(
                    "By Team"
                )
            )

            team_headers = [
                "Fantasy Team",
                "Round",
                "Pick in Round",
                "Overall Pick",
                "Player Name",
                "Position",
                "NFL Team",
                "ESPN Player ID",
            ]

            team_sheet.append(
                team_headers
            )

            picks_by_team = {}

            for pick in draft_picks:

                team_id = pick[
                    "espn_team_id"
                ]

                if (
                    team_id
                    not in picks_by_team
                ):
                    picks_by_team[
                        team_id
                    ] = []

                picks_by_team[
                    team_id
                ].append(
                    pick
                )

            for team in draft_order:

                team_id = team[
                    "espn_team_id"
                ]

                team_name = team[
                    "team_name"
                ]

                team_picks = picks_by_team.get(
                    team_id,
                    [],
                )

                for pick in team_picks:

                    team_sheet.append(
                        [
                            team_name,
                            pick[
                                "round_number"
                            ],
                            pick[
                                "pick_in_round"
                            ],
                            pick[
                                "overall_pick"
                            ],
                            pick[
                                "player_name"
                            ],
                            pick[
                                "position"
                            ],
                            pick[
                                "nfl_team"
                            ],
                            pick[
                                "espn_player_id"
                            ],
                        ]
                    )

            self.format_export_worksheet(
                team_sheet
            )

            # =================================================
            # PICK OWNERSHIP
            # =================================================

            ownership_sheet = (
                workbook.create_sheet(
                    "Pick Ownership"
                )
            )

            ownership_headers = [
                "Overall Pick",
                "Round",
                "Pick in Round",
                "Original Team",
                "Original ESPN Team ID",
                "Current Team",
                "Current ESPN Team ID",
                "Traded",
            ]

            ownership_sheet.append(
                ownership_headers
            )

            for ownership in ownership_rows:

                ownership_sheet.append(
                    [
                        ownership[
                            "overall_pick"
                        ],
                        ownership[
                            "round_number"
                        ],
                        ownership[
                            "pick_in_round"
                        ],
                        ownership[
                            "original_team_name"
                        ],
                        ownership[
                            "original_espn_team_id"
                        ],
                        ownership[
                            "current_team_name"
                        ],
                        ownership[
                            "current_espn_team_id"
                        ],
                        (
                            "Yes"
                            if ownership[
                                "traded"
                            ]
                            else "No"
                        ),
                    ]
                )

            self.format_export_worksheet(
                ownership_sheet
            )

            # =================================================
            # TRADE HISTORY
            # =================================================

            trade_sheet = (
                workbook.create_sheet(
                    "Trade History"
                )
            )

            trade_headers = [
                "Trade ID",
                "Pick A",
                "Pick A Previous Owner",
                "Pick A New Owner",
                "Pick B",
                "Pick B Previous Owner",
                "Pick B New Owner",
                "Recorded By Discord User ID",
                "Traded At",
            ]

            trade_sheet.append(
                trade_headers
            )

            for trade in trade_rows:

                trade_sheet.append(
                    [
                        trade[
                            "id"
                        ],
                        trade[
                            "pick_a"
                        ],
                        trade[
                            "pick_a_from_team_name"
                        ],
                        trade[
                            "pick_a_to_team_name"
                        ],
                        trade[
                            "pick_b"
                        ],
                        trade[
                            "pick_b_from_team_name"
                        ],
                        trade[
                            "pick_b_to_team_name"
                        ],
                        trade[
                            "discord_user_id"
                        ],
                        trade[
                            "traded_at"
                        ],
                    ]
                )

            self.format_export_worksheet(
                trade_sheet
            )

            workbook.save(
                excel_path
            )

            csv_attachment = (
                discord.File(
                    csv_path,
                    filename=(
                        "megabowl_draft_results.csv"
                    ),
                )
            )

            excel_attachment = (
                discord.File(
                    excel_path,
                    filename=(
                        "megabowl_draft_results.xlsx"
                    ),
                )
            )

            await interaction.followup.send(
                content=(
                    "✅ Draft export generated.\n\n"
                    "The Excel workbook includes:\n"
                    "• Draft Results\n"
                    "• By Team\n"
                    "• Pick Ownership\n"
                    "• Trade History"
                ),
                files=[
                    excel_attachment,
                    csv_attachment,
                ],
                ephemeral=True,
            )

        except Exception as error:
            print(
                (
                    "Draft Export Error: "
                    f"{error}"
                )
            )

            await interaction.followup.send(
                (
                    "Something went wrong while "
                    "exporting the draft."
                ),
                ephemeral=True,
            )

    # =========================================================
    # END DRAFT
    # =========================================================

    @app_commands.command(
        name="end-draft",
        description=(
            "End the current fantasy draft."
        ),
    )
    async def end_draft_command(
        self,
        interaction: discord.Interaction,
    ):
        if not self.has_draft_admin_role(
            interaction
        ):
            await interaction.response.send_message(
                (
                    "❌ Only members with the "
                    "**Commissioner** or "
                    "**Developer** role can "
                    "end the draft."
                ),
                ephemeral=True,
            )
            return

        state = get_draft_state()

        if not state["active"]:
            await interaction.response.send_message(
                (
                    "❌ There is no active draft."
                ),
                ephemeral=True,
            )
            return

        picks = (
            get_all_draft_picks()
        )

        view = (
            EndDraftConfirmationView(
                draft_cog=self,
                requesting_user_id=(
                    interaction.user.id
                ),
            )
        )

        await interaction.response.send_message(
            (
                "⚠️ **Are you sure you want "
                "to end the draft?**\n\n"
                f"Completed picks: "
                f"**{len(picks)}**\n"
                "This will stop new picks "
                "from being made."
            ),
            view=view,
            ephemeral=True,
        )


# =============================================================
# COG SETUP
# =============================================================

async def setup(
    bot,
):
    await bot.add_cog(
        DraftCog(
            bot
        )
    )